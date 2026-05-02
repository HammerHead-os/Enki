#!/usr/bin/env python3
"""Generate a single Excel workbook with all datasets and sources."""
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style helpers ──
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
SRC_FONT = Font(italic=True, color="2980B9", size=10)
NOTE_FONT = Font(italic=True, color="7F8C8D", size=9)
TITLE_FONT = Font(bold=True, size=14, color="C0392B")
SUBTITLE_FONT = Font(bold=True, size=11, color="2C3E50")
LINK_FONT = Font(underline="single", color="2980B9", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D5D8DC")
)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, 55)

def add_source_block(ws, row, source_text, url=""):
    ws.cell(row=row, column=1, value="SOURCE:").font = Font(bold=True, size=9, color="C0392B")
    c = ws.cell(row=row, column=2, value=source_text)
    c.font = SRC_FONT
    if url:
        ws.cell(row=row + 1, column=1, value="URL:").font = Font(bold=True, size=9, color="C0392B")
        uc = ws.cell(row=row + 1, column=2, value=url)
        uc.font = LINK_FONT
        uc.hyperlink = url
    return row + 3

# ═══════════════════════════════════════════════════════════
# SHEET 1: SOURCES & OVERVIEW
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Sources & Overview"
ws.sheet_properties.tabColor = "C0392B"

ws.cell(row=1, column=1, value="Internet Association — IL Lobbying Data Compendium").font = TITLE_FONT
ws.cell(row=2, column=1, value="All datasets used in the interactive sandbox visualization, with full source attribution.").font = NOTE_FONT
ws.merge_cells("A1:F1")
ws.merge_cells("A2:F2")

r = 4
ws.cell(row=r, column=1, value="DATASET SOURCES").font = SUBTITLE_FONT
r += 2

sources = [
    ("Sheet", "Dataset", "Source", "URL", "Notes"),
    ("IL Lobbying Positions", "Internet Association lobbying position records in Illinois (2016–2021)", "Original CSV dataset (positions-lobbyistFirmName-internet.csv)", "N/A — provided dataset", "73 records of testimony positions on IL state bills"),
    ("IA Federal Spending", "Internet Association annual federal lobbying expenditure (2013–2021)", "OpenSecrets (Center for Responsive Politics) — federal lobbying disclosures", "https://www.opensecrets.org/federal-lobbying/clients/summary?cycle=2021&id=D000067668", "IA reported $13.9M total 2013–2020. Also referenced via BigTechWiki."),
    ("Internet Sector Spending", "Total US internet sector lobbying spending by year + top company spenders", "Statista / OpenSecrets", "https://www.statista.com/statistics/1035942/amount-spent-internet-sector-lobbying-usa/", "Sector-wide context for IA's share of internet industry lobbying."),
    ("Bill Outcomes", "Legislative outcome (enacted/failed) for each unique bill in the dataset", "Illinois General Assembly (ilga.gov) — bill status lookup", "https://www.ilga.gov/legislation/", "26 unique bills tracked. 6 enacted, 20 failed. 73% IA alignment rate."),
    ("IA Members", "Internet Association member companies, key events, founding & later members", "Wikipedia, VentureBeat, MarketWatch, Engadget, CNET", "https://en.wikipedia.org/wiki/Internet_Association", "~40 members at peak. Founded 2012, dissolved Dec 2021."),
    ("Privacy Legislation Context", "National & IL privacy/tech legislation milestones (2008–2021)", "NCSL, IAPP, Wikipedia — compiled from multiple public sources", "https://www.ncsl.org/technology-and-communication/2021-consumer-data-privacy-legislation", "Key events: BIPA, GDPR, Cambridge Analytica, CCPA, Wayfair ruling."),
    ("US State Privacy Landscape", "State biometric laws, comprehensive privacy laws, online sales tax, IA offices, member HQs", "NCSL, IAPP, Wikipedia — compiled from multiple public sources", "https://iapp.org/resources/article/us-state-privacy-legislation-tracker", "Provides geographic context for why IL was a lobbying battleground."),
]

headers = sources[0]
for ci, h in enumerate(headers, 1):
    ws.cell(row=r, column=ci, value=h)
style_header(ws, r, len(headers))
r += 1

for row_data in sources[1:]:
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.border = THIN_BORDER
        if ci == 4 and val.startswith("http"):
            cell.font = LINK_FONT
            cell.hyperlink = val
    r += 1

r += 2
ws.cell(row=r, column=1, value="GENERATED").font = Font(bold=True, size=9, color="C0392B")
ws.cell(row=r, column=2, value="All data compiled for academic/research visualization purposes. Content rephrased for licensing compliance.").font = NOTE_FONT
auto_width(ws)

# ═══════════════════════════════════════════════════════════
# SHEET 2: IL LOBBYING POSITIONS (original CSV)
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("IL Lobbying Positions")
ws2.sheet_properties.tabColor = "E74C3C"

r = 1
ws2.cell(row=r, column=1, value="IL Lobbying Positions — Internet Association").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws2, r, "Original dataset: positions-lobbyistFirmName-internet.csv", "")

# Parse CSV
with open("positions-lobbyistFirmName-internet.csv", "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    csv_headers = next(reader)
    csv_headers = [h.strip() for h in csv_headers if h.strip()]
    csv_rows = []
    for row in reader:
        cleaned = [c.strip() for c in row]
        if cleaned and cleaned[0]:
            csv_rows.append(cleaned[:len(csv_headers)])

# Write headers
for ci, h in enumerate(csv_headers, 1):
    ws2.cell(row=r, column=ci, value=h)
style_header(ws2, r, len(csv_headers))
r += 1

# Write data
for row_data in csv_rows:
    for ci, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=ci, value=val)
        cell.border = THIN_BORDER
    r += 1

auto_width(ws2)

# ═══════════════════════════════════════════════════════════
# SHEET 3: IA FEDERAL SPENDING
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("IA Federal Spending")
ws3.sheet_properties.tabColor = "F39C12"

with open("data/ia-federal-lobbying.json") as f:
    ia_data = json.load(f)

r = 1
ws3.cell(row=r, column=1, value="Internet Association — Federal Lobbying Spending").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws3, r,
    "OpenSecrets (Center for Responsive Politics) — federal lobbying disclosures",
    "https://www.opensecrets.org/federal-lobbying/clients/summary?cycle=2021&id=D000067668")

headers = ["Year", "Amount ($)", "Amount ($M)", "Registered Lobbyists"]
for ci, h in enumerate(headers, 1):
    ws3.cell(row=r, column=ci, value=h)
style_header(ws3, r, len(headers))
r += 1

for entry in ia_data["annual_spending"]:
    ws3.cell(row=r, column=1, value=entry["year"]).border = THIN_BORDER
    ws3.cell(row=r, column=2, value=entry["amount"]).border = THIN_BORDER
    ws3.cell(row=r, column=2).number_format = '#,##0'
    ws3.cell(row=r, column=3, value=round(entry["amount"] / 1e6, 2)).border = THIN_BORDER
    ws3.cell(row=r, column=3).number_format = '#,##0.00'
    ws3.cell(row=r, column=4, value=entry["lobbyists"]).border = THIN_BORDER
    r += 1

r += 2
ws3.cell(row=r, column=1, value="Top Federal Issues Lobbied").font = SUBTITLE_FONT
r += 1
ws3.cell(row=r, column=1, value="Rank")
ws3.cell(row=r, column=2, value="Issue")
style_header(ws3, r, 2)
r += 1
for i, issue in enumerate(ia_data["top_issues_lobbied"], 1):
    ws3.cell(row=r, column=1, value=i).border = THIN_BORDER
    ws3.cell(row=r, column=2, value=issue).border = THIN_BORDER
    r += 1

auto_width(ws3)

# ═══════════════════════════════════════════════════════════
# SHEET 4: INTERNET SECTOR SPENDING
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Internet Sector Spending")
ws4.sheet_properties.tabColor = "3498DB"

with open("data/internet-sector-lobbying.json") as f:
    sec_data = json.load(f)

r = 1
ws4.cell(row=r, column=1, value="US Internet Sector — Total Lobbying Spending").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws4, r,
    "Statista / OpenSecrets",
    "https://www.statista.com/statistics/1035942/amount-spent-internet-sector-lobbying-usa/")

headers = ["Year", "Total Sector Spend ($M)"]
for ci, h in enumerate(headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(headers))
r += 1

for entry in sec_data["sector_spending_millions"]:
    ws4.cell(row=r, column=1, value=entry["year"]).border = THIN_BORDER
    ws4.cell(row=r, column=2, value=entry["amount"]).border = THIN_BORDER
    ws4.cell(row=r, column=2).number_format = '#,##0.0'
    r += 1

r += 2
ws4.cell(row=r, column=1, value="Top Internet Company Lobbying Spenders (2020)").font = SUBTITLE_FONT
r += 1
headers = ["Company", "Lobbying Spend ($M)", "IA Member?"]
for ci, h in enumerate(headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(headers))
r += 1

for entry in sorted(sec_data["top_spenders_2020"], key=lambda x: -x["amount_millions"]):
    ws4.cell(row=r, column=1, value=entry["company"]).border = THIN_BORDER
    ws4.cell(row=r, column=2, value=entry["amount_millions"]).border = THIN_BORDER
    ws4.cell(row=r, column=2).number_format = '#,##0.00'
    ws4.cell(row=r, column=3, value="Yes" if entry["ia_member"] else "No").border = THIN_BORDER
    r += 1

auto_width(ws4)

# ═══════════════════════════════════════════════════════════
# SHEET 5: BILL OUTCOMES
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Bill Outcomes")
ws5.sheet_properties.tabColor = "27AE60"

with open("data/bill-outcomes.json") as f:
    bill_data = json.load(f)

r = 1
ws5.cell(row=r, column=1, value="Bill Outcomes — IL Legislation in Dataset").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws5, r,
    "Illinois General Assembly (ilga.gov) — bill status lookup",
    "https://www.ilga.gov/legislation/")

headers = ["Bill", "Session", "Title", "IA Position", "Status", "Outcome Matches IA?", "Public Act #"]
for ci, h in enumerate(headers, 1):
    ws5.cell(row=r, column=ci, value=h)
style_header(ws5, r, len(headers))
r += 1

for b in bill_data["bills"]:
    ws5.cell(row=r, column=1, value=b["bill"]).border = THIN_BORDER
    ws5.cell(row=r, column=2, value=b["session"]).border = THIN_BORDER
    ws5.cell(row=r, column=3, value=b["title"]).border = THIN_BORDER
    ws5.cell(row=r, column=4, value="Oppose" if b["ia_position"] == "OPP" else "Support").border = THIN_BORDER
    ws5.cell(row=r, column=4).font = Font(color="C0392B" if b["ia_position"] == "OPP" else "27AE60", bold=True)
    ws5.cell(row=r, column=5, value=b["status"].capitalize()).border = THIN_BORDER
    ws5.cell(row=r, column=5).font = Font(color="F39C12" if b["status"] == "enacted" else "95A5A6", bold=b["status"] == "enacted")
    ws5.cell(row=r, column=6, value="Yes" if b["outcome_matches_ia"] else "No").border = THIN_BORDER
    ws5.cell(row=r, column=6).font = Font(color="27AE60" if b["outcome_matches_ia"] else "C0392B")
    ws5.cell(row=r, column=7, value=b.get("public_act", "")).border = THIN_BORDER
    r += 1

r += 2
ws5.cell(row=r, column=1, value="Summary").font = SUBTITLE_FONT
r += 1
summary = bill_data["summary"]
for key, val in summary.items():
    ws5.cell(row=r, column=1, value=key.replace("_", " ").title()).border = THIN_BORDER
    ws5.cell(row=r, column=2, value=val).border = THIN_BORDER
    r += 1

auto_width(ws5)

# ═══════════════════════════════════════════════════════════
# SHEET 6: IA MEMBERS
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("IA Members")
ws6.sheet_properties.tabColor = "8E44AD"

with open("data/ia-members.json") as f:
    mem_data = json.load(f)

r = 1
ws6.cell(row=r, column=1, value="Internet Association — Member Companies").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws6, r,
    "Wikipedia, VentureBeat, MarketWatch, Engadget, CNET — compiled from multiple public sources",
    "https://en.wikipedia.org/wiki/Internet_Association")

headers = ["Company", "Membership Type", "Status"]
for ci, h in enumerate(headers, 1):
    ws6.cell(row=r, column=ci, value=h)
style_header(ws6, r, len(headers))
r += 1

for m in mem_data["founding_members_2012"]:
    departed = m in mem_data["departed_before_shutdown"]
    ws6.cell(row=r, column=1, value=m).border = THIN_BORDER
    ws6.cell(row=r, column=2, value="Founding Member (2012)").border = THIN_BORDER
    ws6.cell(row=r, column=3, value="Departed before shutdown" if departed else "Active until shutdown").border = THIN_BORDER
    if departed:
        ws6.cell(row=r, column=3).font = Font(color="C0392B")
    r += 1

for m in mem_data["additional_members_by_2018"]:
    departed = m in mem_data["departed_before_shutdown"]
    ws6.cell(row=r, column=1, value=m).border = THIN_BORDER
    ws6.cell(row=r, column=2, value="Joined by 2018").border = THIN_BORDER
    ws6.cell(row=r, column=3, value="Departed before shutdown" if departed else "Active until shutdown").border = THIN_BORDER
    if departed:
        ws6.cell(row=r, column=3).font = Font(color="C0392B")
    r += 1

r += 2
ws6.cell(row=r, column=1, value="Key Events").font = SUBTITLE_FONT
r += 1
headers = ["Date", "Event"]
for ci, h in enumerate(headers, 1):
    ws6.cell(row=r, column=ci, value=h)
style_header(ws6, r, len(headers))
r += 1

for e in mem_data["key_events"]:
    ws6.cell(row=r, column=1, value=e["date"]).border = THIN_BORDER
    ws6.cell(row=r, column=2, value=e["event"]).border = THIN_BORDER
    r += 1

auto_width(ws6)

# ═══════════════════════════════════════════════════════════
# SHEET 7: PRIVACY LEGISLATION CONTEXT
# ═══════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Privacy Legislation Context")
ws7.sheet_properties.tabColor = "1ABC9C"

with open("data/privacy-legislation-context.json") as f:
    ctx_data = json.load(f)

r = 1
ws7.cell(row=r, column=1, value="Privacy & Tech Legislation Milestones").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws7, r,
    "NCSL, IAPP, Wikipedia — compiled from multiple public sources",
    "https://www.ncsl.org/technology-and-communication/2021-consumer-data-privacy-legislation")

ws7.cell(row=r, column=1, value="NATIONAL MILESTONES").font = SUBTITLE_FONT
r += 1
headers = ["Date", "Event", "Relevance"]
for ci, h in enumerate(headers, 1):
    ws7.cell(row=r, column=ci, value=h)
style_header(ws7, r, len(headers))
r += 1

for m in ctx_data["national_milestones"]:
    ws7.cell(row=r, column=1, value=m["date"]).border = THIN_BORDER
    ws7.cell(row=r, column=2, value=m["event"]).border = THIN_BORDER
    ws7.cell(row=r, column=3, value=m["relevance"].upper()).border = THIN_BORDER
    if m["relevance"] == "high":
        ws7.cell(row=r, column=3).font = Font(color="C0392B", bold=True)
    r += 1

r += 2
ws7.cell(row=r, column=1, value="ILLINOIS-SPECIFIC MILESTONES").font = SUBTITLE_FONT
r += 1
headers = ["Date", "Event", "Category"]
for ci, h in enumerate(headers, 1):
    ws7.cell(row=r, column=ci, value=h)
style_header(ws7, r, len(headers))
r += 1

for m in ctx_data["illinois_specific"]:
    ws7.cell(row=r, column=1, value=m["date"]).border = THIN_BORDER
    ws7.cell(row=r, column=2, value=m["event"]).border = THIN_BORDER
    ws7.cell(row=r, column=3, value=m["category"].title()).border = THIN_BORDER
    r += 1

r += 2
ws7.cell(row=r, column=1, value="BILL TOPIC CATEGORIES IN DATASET").font = SUBTITLE_FONT
r += 1
headers = ["Category", "Bill Titles"]
for ci, h in enumerate(headers, 1):
    ws7.cell(row=r, column=ci, value=h)
style_header(ws7, r, len(headers))
r += 1

for cat, titles in ctx_data["bill_topics_in_dataset"].items():
    ws7.cell(row=r, column=1, value=cat.replace("_", " ").title()).border = THIN_BORDER
    ws7.cell(row=r, column=2, value=", ".join(titles)).border = THIN_BORDER
    r += 1

auto_width(ws7)

# ═══════════════════════════════════════════════════════════
# SHEET 8: US STATE PRIVACY LANDSCAPE
# ═══════════════════════════════════════════════════════════
ws8 = wb.create_sheet("US State Privacy Landscape")
ws8.sheet_properties.tabColor = "2980B9"

with open("data/us-state-privacy-laws.json") as f:
    us_data = json.load(f)

r = 1
ws8.cell(row=r, column=1, value="US State Privacy & Tech Regulation Landscape").font = SUBTITLE_FONT
r += 1
r = add_source_block(ws8, r,
    "NCSL, IAPP, Wikipedia — compiled from multiple public sources",
    "https://iapp.org/resources/article/us-state-privacy-legislation-tracker")

ws8.cell(row=r, column=1, value="STATES WITH BIOMETRIC PRIVACY LAWS").font = SUBTITLE_FONT
r += 1
headers = ["State", "Law", "Year Enacted", "Private Right of Action", "Strictness"]
for ci, h in enumerate(headers, 1):
    ws8.cell(row=r, column=ci, value=h)
style_header(ws8, r, len(headers))
r += 1

for state, info in us_data["states_with_biometric_laws"].items():
    ws8.cell(row=r, column=1, value=state).border = THIN_BORDER
    ws8.cell(row=r, column=2, value=info["law"]).border = THIN_BORDER
    ws8.cell(row=r, column=3, value=info["year"]).border = THIN_BORDER
    ws8.cell(row=r, column=4, value="Yes" if info["private_right_of_action"] else "No").border = THIN_BORDER
    ws8.cell(row=r, column=5, value=info["strictness"].replace("_", " ").title()).border = THIN_BORDER
    r += 1

r += 2
ws8.cell(row=r, column=1, value="STATES WITH COMPREHENSIVE PRIVACY LAWS (by 2021)").font = SUBTITLE_FONT
r += 1
for s in us_data["states_with_comprehensive_privacy_by_2021"]:
    ws8.cell(row=r, column=1, value=s).border = THIN_BORDER
    r += 1

r += 2
ws8.cell(row=r, column=1, value="IA OFFICE LOCATIONS").font = SUBTITLE_FONT
r += 1
headers = ["State", "City", "Type", "Latitude", "Longitude"]
for ci, h in enumerate(headers, 1):
    ws8.cell(row=r, column=ci, value=h)
style_header(ws8, r, len(headers))
r += 1

for state, info in us_data["ia_office_locations"].items():
    ws8.cell(row=r, column=1, value=state).border = THIN_BORDER
    ws8.cell(row=r, column=2, value=info["city"]).border = THIN_BORDER
    ws8.cell(row=r, column=3, value=info["type"].replace("_", " ").title()).border = THIN_BORDER
    ws8.cell(row=r, column=4, value=info["lat"]).border = THIN_BORDER
    ws8.cell(row=r, column=5, value=info["lon"]).border = THIN_BORDER
    r += 1

r += 2
ws8.cell(row=r, column=1, value="IA MEMBER COMPANY HQ LOCATIONS").font = SUBTITLE_FONT
r += 1
headers = ["State", "Companies", "Count"]
for ci, h in enumerate(headers, 1):
    ws8.cell(row=r, column=ci, value=h)
style_header(ws8, r, len(headers))
r += 1

for state, companies in sorted(us_data["ia_member_hq_states"].items(), key=lambda x: -len(x[1])):
    ws8.cell(row=r, column=1, value=state).border = THIN_BORDER
    ws8.cell(row=r, column=2, value=", ".join(companies)).border = THIN_BORDER
    ws8.cell(row=r, column=3, value=len(companies)).border = THIN_BORDER
    r += 1

r += 2
ws8.cell(row=r, column=1, value="STATE PRIVACY BILL COUNTS (2019)").font = SUBTITLE_FONT
r += 1
headers = ["State", "Privacy Bills Introduced"]
for ci, h in enumerate(headers, 1):
    ws8.cell(row=r, column=ci, value=h)
style_header(ws8, r, len(headers))
r += 1

for state, count in sorted(us_data["state_privacy_bill_counts_2019"].items(), key=lambda x: -x[1]):
    ws8.cell(row=r, column=1, value=state).border = THIN_BORDER
    ws8.cell(row=r, column=2, value=count).border = THIN_BORDER
    if state == "IL":
        ws8.cell(row=r, column=1).font = Font(bold=True, color="C0392B")
        ws8.cell(row=r, column=2).font = Font(bold=True, color="C0392B")
    r += 1

r += 2
ws8.cell(row=r, column=1, value="STATES WITH ONLINE SALES TAX (Post-Wayfair)").font = SUBTITLE_FONT
r += 1
ws8.cell(row=r, column=1, value=", ".join(sorted(us_data["states_with_online_sales_tax_post_wayfair"]))).border = THIN_BORDER
ws8.cell(row=r + 1, column=1, value=f"Total: {len(us_data['states_with_online_sales_tax_post_wayfair'])} states").font = NOTE_FONT

auto_width(ws8)

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
output_path = "IA_Lobbying_Data_Compendium.xlsx"
wb.save(output_path)
print(f"✅ Excel workbook saved: {output_path}")
print(f"   Sheets: {', '.join(wb.sheetnames)}")
